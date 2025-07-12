from flask import Flask, send_from_directory, request, jsonify, send_file
import os
import openpyxl
from openpyxl import Workbook
from flask_cors import CORS
from datetime import datetime

# Set up Flask to serve from the 'dist' folder
app = Flask(__name__, static_folder='dist', static_url_path='')
CORS(app)

EXCEL_FILE = "spin_responses.xlsx"
PASSWORD = "ecoil123"

# ✅ Serve index.html for root
@app.route('/')
def serve_index():
    return send_from_directory(app.static_folder, 'index.html')

# ✅ Serve admin.html for /admin route
@app.route('/admin')
def serve_admin():
    return send_from_directory(app.static_folder, 'admin.html')

# ✅ Serve static assets (JS, CSS, etc.)
@app.route('/<path:path>')
def serve_static(path):
    return send_from_directory(app.static_folder, path)

# ✅ Save spin result to Excel
@app.route('/submit_spin', methods=['POST'])
def submit_spin():
    data = request.json
    name = data.get('name')
    phone = data.get('phone')
    role = data.get('role')
    prize = data.get('prize')
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Phone", "Role", "Prize", "Timestamp"])
        wb.save(EXCEL_FILE)

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([name, phone, role, prize, timestamp])
    wb.save(EXCEL_FILE)

    return jsonify({"message": "Spin result saved successfully."}), 200

# ✅ Download Excel
@app.route('/download_excel', methods=['GET'])
def download_excel():
    pwd = request.args.get('password')
    if pwd != PASSWORD:
        return jsonify({"error": "Unauthorized"}), 403
    return send_file(EXCEL_FILE, as_attachment=True)

# ✅ Clear Excel
@app.route('/clear_excel', methods=['POST'])
def clear_excel():
    pwd = request.json.get('password')
    if pwd != PASSWORD:
        return jsonify({"error": "Unauthorized"}), 403

    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Phone", "Role", "Prize", "Timestamp"])
    wb.save(EXCEL_FILE)
    return jsonify({"message": "Excel data cleared successfully."}), 200

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=10000)
